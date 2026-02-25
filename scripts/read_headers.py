import openpyxl
from pathlib import Path

EXCEL = Path(__file__).parent.parent / 'Docs' / '2026-Theo Doi Tien Do Ban Hanh VBQPPL.xlsx'
out = []

wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
out.append(f'Sheets: {wb.sheetnames}\n')

for sname in wb.sheetnames:
    ws = wb[sname]
    row1 = None
    row2 = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: row1 = list(row)
        if i == 1: row2 = list(row)
        if i >= 1: break

    out.append(f'=== {sname} ===')
    if row1:
        for ci, c in enumerate(row1):
            val = str(c).strip() if c is not None else ''
            if val:
                out.append(f'  R1-C{ci}: {repr(val[:80])}')
    if row2:
        out.append('  -- Row2 sub-headers --')
        for ci, c in enumerate(row2):
            val = str(c).strip() if c is not None else ''
            if val:
                out.append(f'  R2-C{ci}: {repr(val[:80])}')
    out.append('')

result = '\n'.join(out)
print(result)
out_path = Path(__file__).parent.parent / 'scripts' / 'headers_out.txt'
out_path.write_text(result, encoding='utf-8')
print(f'\nWritten to {out_path}')
