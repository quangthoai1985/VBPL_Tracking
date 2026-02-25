"""
Script Import dữ liệu từ Excel vào Supabase VBPL
Sở Tư pháp Tỉnh An Giang - Năm 2026

Cài đặt: pip install openpyxl supabase python-dotenv
Chạy:    python scripts/import_excel.py
"""

import os
import openpyxl
from pathlib import Path
from supabase import create_client, Client
from dotenv import load_dotenv

# Luôn tìm .env.local tại root project, bất kể chạy script từ đâu
PROJECT_ROOT = Path(__file__).parent.parent
load_dotenv(PROJECT_ROOT / '.env.local')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY')   # Dùng service_role để bypass RLS

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError(
        "❌ Không tìm thấy SUPABASE_URL hoặc SUPABASE_SERVICE_ROLE_KEY.\n"
        f"   Kiểm tra file: {PROJECT_ROOT / '.env.local'}\n"
        "   Cần có:\n"
        "     NEXT_PUBLIC_SUPABASE_URL=https://xxx.supabase.co\n"
        "     SUPABASE_SERVICE_ROLE_KEY=eyJ...  (lấy từ Supabase Dashboard > Settings > API)"
    )

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

EXCEL_FILE = PROJECT_ROOT / 'Docs' / '2026-Theo Doi Tien Do Ban Hanh VBQPPL.xlsx'

# Thứ tự đúng theo sheet Excel
SHEET_MAP = [
    ('NQ can xu ly',      {'doc_type': 'NQ',          'status': 'can_xu_ly'}),
    ('QD UBND can xu ly', {'doc_type': 'QD_UBND',     'status': 'can_xu_ly'}),
    ('QD CT.UBND',        {'doc_type': 'QD_CT_UBND',  'status': 'can_xu_ly'}),
    ('NQ HDND da xu ly',  {'doc_type': 'NQ',          'status': 'da_xu_ly'}),
    ('QD UBND da xu ly',  {'doc_type': 'QD_UBND',     'status': 'da_xu_ly'}),
]

# Cache agencies
agency_cache: dict[str, int] = {}


def get_or_create_agency(name: str) -> int | None:
    """Lấy hoặc tạo agency_id theo tên."""
    if not name or str(name).strip() == '':
        return None
    name = str(name).strip()
    if name in agency_cache:
        return agency_cache[name]
    # Tìm trong DB
    res = supabase.table('agencies').select('id').eq('name', name).execute()
    if res.data:
        agency_cache[name] = res.data[0]['id']
        return res.data[0]['id']
    # Tạo mới
    res = supabase.table('agencies').insert({'name': name}).execute()
    if res.data:
        agency_cache[name] = res.data[0]['id']
        print(f"  + Tạo agency mới: {name}")
        return res.data[0]['id']
    return None


def safe_str(val) -> str | None:
    """Chuyển giá trị cell thành string, bỏ qua None/rỗng."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' else None


def to_int(val) -> int:
    """Chuyển giá trị cell thành int >= 0."""
    if val is None:
        return 0
    try:
        n = int(float(str(val)))
        return max(0, n)
    except (ValueError, TypeError):
        return 0


def find_col(headers: list, *keywords: str) -> int | None:
    """Tìm index cột theo keyword."""
    for i, h in enumerate(headers):
        if h and any(kw.lower() in str(h).lower() for kw in keywords):
            return i
    return None


def parse_sheet(ws, doc_type: str, status: str) -> list[dict]:
    """Parse một sheet Excel thành danh sách records."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Row 1 là header chính, Row 2 là sub-header
    header_row1 = [str(c).strip() if c else '' for c in rows[0]]
    # Row 2: col[3]=Thay thế, col[4]=Bãi bỏ, col[5]=Ban hành mới, col[6]=Chưa xác định

    # Tìm các cột
    col_stt      = find_col(header_row1, 'STT', 'stt')
    col_name     = find_col(header_row1, 'Tên gọi', 'TÊN GỌI')
    col_agency   = find_col(header_row1, 'Cơ quan soạn')
    col_handler  = find_col(header_row1, 'Người xử lý')
    # Cột "Hình thức xử lý" là cột gốc, các sub-cột offset từ đó
    col_htxl     = find_col(header_row1, 'Hình thức xử lý')
    base         = col_htxl if col_htxl is not None else 3  # mặc định col 3

    col_reg_agency    = find_col(header_row1, 'đăng ký xây dựng NQ của cơ quan', 'VB đăng ký xây dựng', 'đăng ký xây dựng')
    col_reg_ubnd      = find_col(header_row1, 'đăng ký xây dựng NQ của UBND')
    col_approval      = find_col(header_row1, 'Ý kiến chấp thuận')
    col_expected      = find_col(header_row1, 'dự kiến trình ban hành', 'Ngày dự kiến')
    col_feedback      = find_col(header_row1, 'lấy ý kiến góp ý')
    col_appraisal     = find_col(header_row1, 'Sở Tư pháp thẩm định', 'gửi Sở Tư pháp')
    col_sub_ubnd      = find_col(header_row1, 'trình UBND tỉnh')
    col_sub_hdnd      = find_col(header_row1, 'UBND tỉnh trình HĐND')
    col_issuance      = find_col(header_row1, 'Số, trích yếu', 'Số, ngày', 'ban hành VBQPPL')
    col_proc_time     = find_col(header_row1, 'Thời gian xử lý')
    col_notes         = find_col(header_row1, 'Ghi chú')

    def gv(row_vals, col) -> str | None:
        if col is None or col >= len(row_vals):
            return None
        return safe_str(row_vals[col])

    def gi(row_vals, col) -> int:
        if col is None or col >= len(row_vals):
            return 0
        return to_int(row_vals[col])

    records = []
    stt_n = 0

    for row_vals in rows[2:]:
        # Bỏ qua dòng rỗng
        name_val = gv(row_vals, col_name)
        if not name_val:
            continue

        stt_n += 1

        # STT
        stt_val = stt_n
        if col_stt is not None and col_stt < len(row_vals):
            try:
                stt_val = int(float(str(row_vals[col_stt])))
            except (ValueError, TypeError):
                pass

        # Cơ quan soạn thảo
        agency_name = gv(row_vals, col_agency)
        agency_id = get_or_create_agency(agency_name) if agency_name else None

        # ---  Hình thức xử lý: đọc 4 số lượng VB ---
        # base+0 = Thay thế, +1 = Bãi bỏ, +2 = Ban hành mới, +3 = Chưa xác định
        count_thay_the      = gi(row_vals, base)
        count_bai_bo        = gi(row_vals, base + 1)
        count_ban_hanh_moi  = gi(row_vals, base + 2)
        count_chua_xac_dinh = gi(row_vals, base + 3)

        # Xác định processing_form chính
        counts = [count_thay_the, count_bai_bo, count_ban_hanh_moi, count_chua_xac_dinh]
        form_keys = ['thay_the', 'bai_bo', 'ban_hanh_moi', 'chua_xac_dinh']
        non_zero = [(v, k) for v, k in zip(counts, form_keys) if v > 0]
        if len(non_zero) == 1:
            processing_form = non_zero[0][1]
        elif len(non_zero) > 1:
            processing_form = max(non_zero, key=lambda x: x[0])[1]
        else:
            processing_form = None

        record = {
            'doc_type':             doc_type,
            'status':               status,
            'stt':                  stt_val,
            'name':                 name_val,
            'agency_id':            agency_id,
            'handler_name':         gv(row_vals, col_handler),
            'processing_form':      processing_form,
            'count_thay_the':       count_thay_the,
            'count_bai_bo':         count_bai_bo,
            'count_ban_hanh_moi':   count_ban_hanh_moi,
            'count_chua_xac_dinh':  count_chua_xac_dinh,
            'reg_doc_agency':       gv(row_vals, col_reg_agency),
            'reg_doc_ubnd':         gv(row_vals, col_reg_ubnd),
            'approval_hdnd':        gv(row_vals, col_approval),
            'expected_date':        gv(row_vals, col_expected),
            'feedback_sent':        gv(row_vals, col_feedback),
            'appraisal_sent':       gv(row_vals, col_appraisal),
            'submitted_ubnd':       gv(row_vals, col_sub_ubnd),
            'submitted_hdnd':       gv(row_vals, col_sub_hdnd),
            'issuance_number':      gv(row_vals, col_issuance),
            'processing_time':      gv(row_vals, col_proc_time),
            'notes':                gv(row_vals, col_notes),
            'year': 2026,
        }
        records.append(record)

    return records


def import_sheet(sheet_name: str, doc_type: str, status: str, wb):
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Không tìm thấy sheet: {sheet_name}")
        return 0

    ws = wb[sheet_name]
    print(f"\n📄 Sheet: [{sheet_name}] → {doc_type} / {status}")
    records = parse_sheet(ws, doc_type, status)
    print(f"   Tổng nhóm VB: {len(records)}")

    # Batch insert (50 records/lần)
    inserted = 0
    BATCH = 50
    for i in range(0, len(records), BATCH):
        batch = records[i:i + BATCH]
        res = supabase.table('documents').insert(batch).execute()
        if res.data:
            inserted += len(res.data)
        else:
            print(f"   ❌ Lỗi batch {i}-{i+BATCH}")

    print(f"   ✅ Đã import thành công: {inserted} nhóm văn bản")
    return inserted


def main():
    print("=" * 60)
    print("IMPORT DỮ LIỆU VBQPPL 2026 → SUPABASE")
    print("Sở Tư pháp Tỉnh An Giang")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f"✅ Đọc file Excel: {EXCEL_FILE}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

    total = 0
    for sheet_name, config in SHEET_MAP:
        count = import_sheet(sheet_name, config['doc_type'], config['status'], wb)
        total += count

    print(f"\n{'=' * 60}")
    print(f"🎉 HOÀN THÀNH! Tổng đã import: {total} nhóm văn bản")
    print(f"   Truy cập Supabase Dashboard để kiểm tra:")
    print(f"   https://supabase.com/dashboard/project/digiwjviiyrhhosegplh")


if __name__ == '__main__':
    main()
