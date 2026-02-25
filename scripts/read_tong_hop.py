import openpyxl

wb = openpyxl.load_workbook('Docs/2026-Theo Doi Tien Do Ban Hanh VBQPPL.xlsx', data_only=True)

with open('scripts/sheet_analysis.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheets: {wb.sheetnames}\n\n")
    
    for sheet_name in ['NQ can xu ly', 'QD UBND can xu ly', 'NQ HDND da xu ly', 'QD UBND da xu ly']:
        if sheet_name not in wb.sheetnames:
            f.write(f"Sheet '{sheet_name}' not found\n")
            continue
        ws = wb[sheet_name]
        f.write(f"=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===\n")
        for r in range(1, min(8, ws.max_row + 1)):
            line = []
            for c in range(1, min(20, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v is not None:
                    line.append(f'C{c}={v}')
            if line:
                f.write(f'  R{r}: {" | ".join(line)}\n')
        f.write("\n")
    
    # Check NQ can xu ly headers and first rows
    ws = wb['NQ can xu ly']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else '' for c in rows[0]]
    f.write(f"NQ can xu ly ALL headers:\n")
    for i, h in enumerate(headers):
        f.write(f"  Col {i}: '{h}'\n")
    
    # Check if agency names map to linh vuc
    f.write("\n\nNQ can xu ly - all agency names (col_agency):\n")
    col_agency = None
    for i, h in enumerate(headers):
        if 'quan' in h.lower() and 'soạn' in h.lower():
            col_agency = i
            break
    if col_agency is not None:
        agencies = set()
        for row_vals in rows[2:]:
            if col_agency < len(row_vals) and row_vals[col_agency]:
                agencies.add(str(row_vals[col_agency]).strip())
        for a in sorted(agencies):
            f.write(f"  - {a}\n")
    
    # Same for QD UBND can xu ly
    ws2 = wb['QD UBND can xu ly']
    rows2 = list(ws2.iter_rows(values_only=True))
    headers2 = [str(c).strip() if c else '' for c in rows2[0]]
    f.write(f"\nQD UBND can xu ly ALL headers:\n")
    for i, h in enumerate(headers2):
        f.write(f"  Col {i}: '{h}'\n")
    
    col_agency2 = None
    for i, h in enumerate(headers2):
        if 'quan' in h.lower() and 'soạn' in h.lower():
            col_agency2 = i
            break
    if col_agency2 is not None:
        agencies2 = set()
        for row_vals in rows2[2:]:
            if col_agency2 < len(row_vals) and row_vals[col_agency2]:
                agencies2.add(str(row_vals[col_agency2]).strip())
        f.write(f"\nQD UBND can xu ly - all agency names:\n")
        for a in sorted(agencies2):
            f.write(f"  - {a}\n")

print("Done! Output in scripts/sheet_analysis.txt")
