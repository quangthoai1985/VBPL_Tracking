import openpyxl
from pathlib import Path

file_path = Path('Docs/2026-Theo Doi Tien Do Ban Hanh VBQPPL.xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

with open('extract_output.txt', 'w', encoding='utf-8') as f:
    if 'Tong Hop Chung' in wb.sheetnames:
        ws = wb['Tong Hop Chung']
        f.write(f"Sheet 'Tong Hop Chung' has {ws.max_row} rows and {ws.max_column} columns.\n")
        
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                row_str = [str(cell)[:100].replace('\n', ' ') if cell is not None else '' for cell in row]
                f.write(f"{row_str}\n")
    else:
        f.write("Sheet 'Tong Hop Chung' not found.\n")
